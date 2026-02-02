from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render

from localesGenerales.forms import ProductoForm
from .models import Inventario

def index(request):
    productos = Inventario.objects.all()
    
    return render(
        request,'index.html',context={'inventario':productos})
    


def error_404_view(request, exception):
    return render(request, '404.html', status=404)


def formulario(request):
    if request.method =='POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect('/localesGenerales')
            
    else:
        form = ProductoForm
        
    return render( request, 'producto_form.html',{'form': form})



from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Inventario

@csrf_exempt
def actualizar_conteo(request):
    if request.method == "POST":
        item_id = request.POST.get("id")
        campo = request.POST.get("campo")
        valor = request.POST.get("valor")

        item = Inventario.objects.get(id=item_id)
        setattr(item, campo, valor)
        item.save()

        return JsonResponse({
            "diferencia": item.diferencia
        })


def to_int(valor):
    try:
        if valor in (None, '', ' '):
            return 0
        return int(valor)
    except (ValueError, TypeError):
        return 0
    
    
import openpyxl
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Inventario

def to_int(valor):
    try:
        return int(valor)
    except (TypeError, ValueError):
        return 0

def importar_excel(request):
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']

        if not archivo.name.endswith('.xlsx'):
            messages.error(request, "El archivo debe ser .xlsx")
            return redirect('importar_excel')

        wb = openpyxl.load_workbook(archivo, data_only=True)
        hoja = wb.active

        creados = 0

        for fila in hoja.iter_rows(min_row=2, values_only=True):
            Inventario.objects.create(
                ubicacion=str(fila[0]).strip() if fila[0] else '',
                cod_ean=str(fila[1]).strip() if fila[1] else '',
                cod_dun=str(fila[2]).strip() if fila[2] else '',
                cod_sistema=str(fila[3]).strip() if fila[3] else '',
                descripcion=str(fila[4]).strip() if fila[4] else '',
                categoria=str(fila[5]).strip() if fila[5] else '',
                conteo_01=to_int(fila[6]),
                conteo_02=to_int(fila[7]),
            )
            creados += 1

        messages.success(
            request,
            f"Excel importado correctamente. Registros creados: {creados}"
        )

        return redirect('index')

    return render(request, 'importar_excel.html')



import openpyxl
from django.http import HttpResponse
from .models import Inventario

def exportar_excel(request):
    # Crear libro y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Cabeceras
    columnas = [
        'Ubicación',
        'Cod EAN',
        'Cod DUN',
        'Cod Sistema',
        'Descripción',
        'Categoría',
        'Conteo 01',
        'Conteo 02',
        'Diferencia',
        'Fecha Creación',
    ]
    ws.append(columnas)

    # Datos
    for item in Inventario.objects.all().order_by('id'):
        ws.append([
            item.ubicacion,
            item.cod_ean,
            item.cod_dun,
            item.cod_sistema,
            item.descripcion,
            item.categoria,
            item.conteo_01,
            item.conteo_02,
            item.diferencia,
            item.creado.strftime('%d-%m-%Y %H:%M'),
        ])

    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=inventario.xlsx'

    wb.save(response)
    return response



from django.shortcuts import render, get_object_or_404, redirect
from .models import Inventario
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse



def editar_inventario(request, pk):
    item = get_object_or_404(Inventario, pk=pk)

    if request.method == 'POST':
        item.ubicacion = request.POST.get('ubicacion')
        item.cod_ean = request.POST.get('cod_ean')
        item.cod_dun = request.POST.get('cod_dun')
        item.cod_sistema = request.POST.get('cod_sistema')
        item.descripcion = request.POST.get('descripcion')
        item.categoria = request.POST.get('categoria')
        item.conteo_01 = int(request.POST.get('conteo_01') or 0)
        item.conteo_02 = int(request.POST.get('conteo_02') or 0)
        item.save()

        return redirect('index')

    return render(request, 'editar_inventario.html', {'item': item})


def eliminar_inventario(request, pk):
    item = get_object_or_404(Inventario, pk=pk)

    if request.method == 'POST':
        item.delete()
        return redirect('index')

    return render(request, 'confirmar_eliminar.html', {'item': item})
